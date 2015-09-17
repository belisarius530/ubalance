from struct import *
import serial
from time import sleep
import re
import openpyxl

def Send_Command(packet, wait = 1):
    ser = serial.Serial('COM9', 115200, timeout = wait)
    trys = 3
    for elem in packet:
        ser.write(pack('B',elem))
    msg = ser.read(50000)
    msg = msg.upper()
    match = re.search("ERROR", msg)
    while (match and(trys>0) ):
        trys -=1
        for elem in packet:
            ser.write(pack('B',elem))
        msg = ser.read(50000)
        msg = msg.upper()
        match = re.search("ERROR", msg)
        print msg
        if match:
            print "Attempt failed, trying again..."
            sleep(2)
    print msg
    ser.close()
    
def Send_Command_Collect_Data(packet, wait = 1, trys= 3, num_pts = 50):
    ser = serial.Serial('COM9', 115200, timeout = wait)
    #trys = 3
    for elem in packet:
        ser.write(pack('B',elem))
    msg = ser.read(50000)
    msg = msg.upper()
    match = re.search("ERROR", msg)
    while (match and(trys>0) ):
        trys -=1
        for elem in packet:
            ser.write(pack('B',elem))
        msg = ser.read(50000)
        msg = msg.upper()
        match = re.search("ERROR", msg)
        print msg
        if match:
            print "Attempt failed, trying again..."
            sleep(2)
    print msg
    ser.close()
    data_match = re.search("ANGLE",msg)
    if data_match:
        data = re.findall(r'[-\d]+\.[\d]+', msg)
        return data
    else:
        return []
        
    
def Construct_Payload(opcode, index, data = []):
    buff = []
    buff.append(pack('B', opcode))
    buff.append(pack('B', index))
    acceptable_data_lengths = [0, 1, 4,16]
    if ( ( len(data) not in acceptable_data_lengths)):
        print "Invalid data length. Data must be in an array of length 1, 4, or 16"
        return []
    if (len(data)>0):
        if(type(data[0])==int):
            for elem in range(len(data)):
                buff.extend(pack('<i', data[elem]))
        elif(type(data[0])==float):
            for elem in range(len(data)):
                buff.extend(pack('<f', data[elem]))
        else:
            print "Incorrect data type. Please use Int or Float types only"
            return []
    return buff

def Generate_Checkbytes(data):
    sum1 = 0
    sum2 = 0
    for elem in data:
        sum1 = (sum1 + ord(elem) )%255
        sum2 = (sum2 + sum1)%255
    c0 = 255 - (sum1+sum2)%255
    c1 = 255 - (sum1 + c0)%255
    return [pack('B', c0), pack('B',c1)]

def Variable_Report(var_index):
    id_l = 9
    id_h = 110
    length = 4
    header = [id_l, id_h, length]
    payload_str = Construct_Payload(opcode = 2, index = var_index)
    if (len(payload_str)<1):
        print "Error: Payload must have length > 0"
        return[]
    payload_str.extend(Generate_Checkbytes(payload_str))
    payload = []
    for elem in payload_str:
        payload.append( ord(elem)%255 )
    Send_Command(header + payload)
    return # header + payload

def Update_Variable(index, data):
    id_l = 9
    id_h = 110
    opcode = 1
    if (len(data)<1):
        print "data length MUST be greater than 1 when updating variables"
        return []
    
    payload_str = Construct_Payload(opcode, index, data)
    if (len(payload_str)<1):
        print "Error: Payload must have length > 0"
        return[]
    payload_str.extend(Generate_Checkbytes(payload_str))
    payload = []
    for elem in payload_str:
        payload.append( ord(elem)%256 )
    length = len(payload)
    header = [id_l, id_h, length]
    Send_Command(header + payload)
    return # header + payload

def Full_Variable_Report():
    id_l = 9
    id_h = 110
    length = 4
    header = [id_l, id_h, length]
    payload_str = Construct_Payload(opcode = 3, index = 0)
    if (len(payload_str)<1):
        print "Error: Payload must have length > 0"
        return[]
    payload_str.extend(Generate_Checkbytes(payload_str))
    payload = []
    for elem in payload_str:
        payload.append( ord(elem)%255 )
    Send_Command(header + payload)
    return # header + payload

def Update_Coefficients():
    id_l = 9
    id_h = 110
    length = 4
    header = [id_l, id_h, length]
    payload_str = Construct_Payload(opcode = 5, index = 0)
    if (len(payload_str)<1):
        print "Error: Payload must have length > 0"
        return[]
    payload_str.extend(Generate_Checkbytes(payload_str))
    payload = []
    for elem in payload_str:
        payload.append( ord(elem)%255 )
    Send_Command(header + payload)
    return # header + payload

def PID(k_prop, k_der, k_int):
    Update_Variable(0, [k_prop])
    sleep(.1)
    Update_Variable(1, [k_der])
    sleep(.1)
    Update_Variable(2, [k_int])
    sleep(.1)
    Update_Coefficients()
    return

def Report_State():
    id_l = 9
    id_h = 110
    length = 4
    header = [id_l, id_h, length]
    payload_str = Construct_Payload(opcode = 7, index = 0)
    if (len(payload_str)<1):
        print "Error: Payload must have length > 0"
        return[]
    payload_str.extend(Generate_Checkbytes(payload_str))
    payload = []
    for elem in payload_str:
        payload.append( ord(elem)%255 )
    Send_Command(header + payload)
    return # header + payload
    
def Initiate_Test():
    id_l = 9
    id_h = 110
    length = 4
    header = [id_l, id_h, length]
    payload_str = Construct_Payload(opcode = 9, index = 0)
    if (len(payload_str)<1):
        print "Error: Payload must have length > 0"
        return[]
    payload_str.extend(Generate_Checkbytes(payload_str))
    payload = []
    for elem in payload_str:
        payload.append( ord(elem)%255 )
    Send_Command(header + payload, 8)
    return # header + payload
    
def Initiate_Test_Return_Excel(filename, num_pts = 50):
    wb = Workbook()
    ws = wb.active
    angle = []
    angular_vel = []
    position = []
    velocity = []
    control_sig = []
    id_l = 9
    id_h = 110
    length = 4
    header = [id_l, id_h, length]
    payload_str = Construct_Payload(opcode = 9, index = 0)
    if (len(payload_str)<1):
        print "Error: Payload must have length > 0"
        return[]
    payload_str.extend(Generate_Checkbytes(payload_str))
    payload = []
    for elem in payload_str:
        payload.append( ord(elem)%255 )
    datalist = Send_Command_Collect_Data(header + payload, 8)
    if(len(datalist)/5==num_pts):
        for elem in range(len(datalist)/5):
            ws.cell(row = elem+1, column = 1).value = elem
            ws.cell(row = elem+1, column = 2).value = float(datalist[elem*4])
            ws.cell(row = elem+1, column = 3).value = float(datalist[(elem*4)+1])
            ws.cell(row = elem+1, column = 4).value = float(datalist[(elem*4)+2])
            ws.cell(row = elem+1, column = 5).value = float(datalist[(elem*4)+3])
            ws.cell(row = elem+1, column = 6).value = float(datalist[(elem*4)+4])
            angle.append(float(datalist[elem*5]))
            angular_vel.append(float(datalist[(elem*5)+1]))
            position.append(float(datalist[(elem*5)+2]))
            velocity.append(float(datalist[(elem*5)+3]))
            control_sig.append(float(datalist[(elem*5)+4]))
        wb.save(filename)
        return [angle, angular_vel, position, velocity, control_sig]
    else:
        print "Error: An incorrect quantity of data was collected."
        return []


def Initiate_Test_Return_Lists(num_pts = 50):

    angle = []
    angular_vel = []
    position = []
    velocity = []
    control_sig = []
    id_l = 9
    id_h = 110
    length = 4
    header = [id_l, id_h, length]
    payload_str = Construct_Payload(opcode = 9, index = 0)
    if (len(payload_str)<1):
        print "Error: Payload must have length > 0"
        return[]
    payload_str.extend(Generate_Checkbytes(payload_str))
    payload = []
    for elem in payload_str:
        payload.append( ord(elem)%255 )
    datalist = Send_Command_Collect_Data(header + payload, 8)
    if(len(datalist)/5==num_pts):
        for elem in range(len(datalist)/5):
            
            angle.append(float(datalist[elem*5]))
            angular_vel.append(float(datalist[(elem*5)+1]))
            position.append(float(datalist[(elem*5)+2]))
            velocity.append(float(datalist[(elem*5)+3]))
            control_sig.append(float(datalist[(elem*5)+4]))
        return [angle, angular_vel, position, velocity, control_sig]
    else:
        print "Error: An incorrect quantity of data was collected."
        return []
    

def Report_Angle():
    id_l = 9
    id_h = 110
    length = 4
    header = [id_l, id_h, length]
    payload_str = Construct_Payload(opcode = 8, index = 0)
    if (len(payload_str)<1):
        print "Error: Payload must have length > 0"
        return[]
    payload_str.extend(Generate_Checkbytes(payload_str))
    payload = []
    for elem in payload_str:
        payload.append( ord(elem)%255 )
    Send_Command(header + payload)
    return

def Report_Control_Law():
    id_l = 9
    id_h = 110
    length = 4
    header = [id_l, id_h, length]
    payload_str = Construct_Payload(opcode = 10, index = 0)
    if (len(payload_str)<1):
        print "Error: Payload must have length > 0"
        return[]
    payload_str.extend(Generate_Checkbytes(payload_str))
    payload = []
    for elem in payload_str:
        payload.append( ord(elem)%255 )
    Send_Command(header + payload)
    return # header + payload
    
def Report_Filter_Law():
    id_l = 9
    id_h = 110
    length = 4
    header = [id_l, id_h, length]
    payload_str = Construct_Payload(opcode = 11, index = 0)
    if (len(payload_str)<1):
        print "Error: Payload must have length > 0"
        return[]
    payload_str.extend(Generate_Checkbytes(payload_str))
    payload = []
    for elem in payload_str:
        payload.append( ord(elem)%255 )
    Send_Command(header + payload)
    return # header + payload